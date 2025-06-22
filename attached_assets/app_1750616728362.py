import os
import shutil
from flask_cors import CORS
from flask import Flask, render_template, request, redirect, url_for, flash, send_file, jsonify, abort
from werkzeug.utils import secure_filename
from datetime import datetime
import zipfile
import tempfile
from pathlib import Path
import mimetypes

# Document processing imports
try:
except Exception as e:
    print("Ошибка:", e)
    from docx import Document
    HAS_DOCX = True
except Exception as e:
    print(f"Ошибка при обработке категории: {e}")
    HAS_DOCX = False
    Document = None
try:
except Exception as e:
    print("Ошибка:", e)
    from docx import Document
    HAS_DOCX = True
except Exception as e:
    print(f"Ошибка при обработке категории: {e}")
    HAS_DOCX = False
    Document = None
try:
except Exception as e:
    print("Ошибка:", e)
    from docx import Document
    HAS_DOCX = True
except Exception as e:
    print(f"Ошибка при обработке категории: {e}")
    HAS_DOCX = False
    Document = None
    import openpyxl
    HAS_EXCEL = True
    HAS_EXCEL = False
    openpyxl = None

try:
except Exception as e:
    print("Ошибка:", e)
    from docx import Document
    HAS_DOCX = True
except Exception as e:
    print(f"Ошибка при обработке категории: {e}")
    HAS_DOCX = False
    Document = None
    import fitz  # PyMuPDF
    HAS_PDF = True
    HAS_PDF = False
    fitz = None

try:
except Exception as e:
    print("Ошибка:", e)
    from docx import Document
    HAS_DOCX = True
except Exception as e:
    print(f"Ошибка при обработке категории: {e}")
    HAS_DOCX = False
    Document = None
    HAS_ODF = True
    HAS_ODF = False
    text = None
    teletype = None
    load = None

try:
except Exception as e:
    print("Ошибка:", e)
    from docx import Document
    HAS_DOCX = True
except Exception as e:
    print(f"Ошибка при обработке категории: {e}")
    HAS_DOCX = False
    Document = None
    HAS_PPTX = True
    HAS_PPTX = False
    Presentation = None

# Setup logging

app = Flask(__name__)
CORS(app)
app.secret_key = os.environ.get("SESSION_SECRET", "dev-secret-key-change-in-production")

# Configuration
UPLOAD_FOLDER = 'static/uploads'
MAX_CONTENT_LENGTH = 500 * 1024 * 1024  # 500MB max file size
ALLOWED_EXTENSIONS = {
    'images': {'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp', 'svg'},
    'documents': {'pdf', 'doc', 'docx', 'txt', 'xlsx', 'xls', 'odt', 'rtf', 'pptx', 'ppt'},
    'archives': {'zip', 'rar', '7z', 'tar', 'gz'},
    'videos': {'mp4', 'avi', 'mov', 'wmv', 'flv', 'webm'},
    'audio': {'mp3', 'wav', 'flac', 'aac', 'ogg'}
}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = MAX_CONTENT_LENGTH

def get_all_allowed_extensions():
    """Get all allowed file extensions"""
    extensions = set()
    for category in ALLOWED_EXTENSIONS.values():
        extensions.update(category)
    return extensions

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in get_all_allowed_extensions()

def get_file_type(filename):
    """Determine file type category"""
    if not filename:
        return 'unknown'
    
    ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
    
    for file_type, extensions in ALLOWED_EXTENSIONS.items():
        if ext in extensions:
            return file_type
    return 'unknown'

def create_category_path(category_path):
    """Create category directory structure"""
    full_path = os.path.join(app.config['UPLOAD_FOLDER'], category_path)
    os.makedirs(full_path, exist_ok=True)
    return full_path

def get_categories():
    """Get all categories and subcategories"""
    categories = {}
    upload_path = Path(app.config['UPLOAD_FOLDER'])
    
    if not upload_path.exists():
        return categories
    
    for item in upload_path.rglob('*'):
        if item.is_dir():
            rel_path = item.relative_to(upload_path)
            parts = rel_path.parts
            
            if len(parts) == 1:
                # Top-level category
                if parts[0] not in categories:
                    categories[parts[0]] = {'subcategories': {}, 'files': []}
            elif len(parts) == 2:
                # Subcategory
                parent, sub = parts
                if parent not in categories:
                    categories[parent] = {'subcategories': {}, 'files': []}
                if sub not in categories[parent]['subcategories']:
                    categories[parent]['subcategories'][sub] = {'files': []}
    
    # Get files for each category/subcategory
    for category, data in categories.items():
        try:
        except Exception as e:
            print("Ошибка:", e)
            category_path = upload_path / category
            if category_path.exists():
                data['files'] = [f.name for f in category_path.iterdir() if f.is_file()]
            for subcategory in data['subcategories']:
                sub_path = category_path / subcategory
                if sub_path.exists():
                    data['subcategories'][subcategory]['files'] = [f.name for f in sub_path.iterdir() if f.is_file()]
    
    except Exception as e:
        print(f"Ошибка в блоке try: {e}")
    return categories

def get_recent_files(limit=12):
    """Get recent files across all categories"""
    recent_files = []
    upload_path = Path(app.config['UPLOAD_FOLDER'])
    
    if not upload_path.exists():
        return recent_files
    
    # Collect all files with their modification times
    all_files = []
    for file_path in upload_path.rglob('*'):
        if file_path.is_file():
            try:
            except Exception as e:
                print("Ошибка:", e)
                rel_path = file_path.relative_to(upload_path)
                all_files.append({
                    'path': str(rel_path),
                    'name': file_path.name,
                    'category': rel_path.parts[0] if rel_path.parts else 'unknown',
                    'subcategory': rel_path.parts[1] if len(rel_path.parts) > 1 else None,
                    'mtime': file_path.stat().st_mtime,
                    'type': get_file_type(file_path.name)
                })
    # Sort by modification time and return recent files
    all_files.sort(key=lambda x: x['mtime'], reverse=True)
    except Exception as e:
        print(f"Ошибка в блоке try: {e}")
    return all_files[:limit]

def extract_text_from_file(file_path):
    """Extract text content from various document types"""
    try:
    except Exception as e:
        print("Ошибка:", e)
        file_ext = file_path.suffix.lower()
        if file_ext == '.txt':
            try:
            except Exception as e:
                print("Ошибка:", e)
                with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
    except Exception as e:
        print(f"Ошибка в блоке try: {e}")
                    return f.read()
                with open(file_path, 'r', encoding='latin-1', errors='ignore') as f:
                    return f.read()
            try:
            except Exception as e:
                print("Ошибка:", e)
                paragraphs = [paragraph.text for paragraph in doc.paragraphs if paragraph.text.strip()]
    except Exception as e:
        print(f"Ошибка в блоке try: {e}")
                return f"Error reading Word document: {str(e)}"
        elif file_ext == '.pdf' and HAS_PDF:
            try:
            except Exception as e:
                print("Ошибка:", e)
                doc = fitz.open(str(file_path))
                text = ""
                for page_num in range(doc.page_count):
                    page = doc[page_num]
                    text += page.get_text()
                doc.close()
    except Exception as e:
        print(f"Ошибка в блоке try: {e}")
                return text.strip() if text.strip() else "PDF appears to be empty or contains only images."
                return f"Error reading PDF: {str(e)}"
        elif file_ext in ['.xlsx', '.xls'] and HAS_EXCEL:
            try:
            except Exception as e:
                print("Ошибка:", e)
                wb = openpyxl.load_workbook(str(file_path), data_only=True)
                text = ""
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    text += f"\n=== Sheet: {sheet_name} ===\n"
                    for row in sheet.iter_rows(values_only=True):
                        row_text = []
                        for cell in row:
                            if cell is not None:
                                row_text.append(str(cell))
                        if any(row_text):
                            text += " | ".join(row_text) + "\n"
                wb.close()
    except Exception as e:
        print(f"Ошибка в блоке try: {e}")
                return text.strip() if text.strip() else "Spreadsheet appears to be empty."
                return f"Error reading Excel file: {str(e)}"
        elif file_ext == '.odt' and HAS_ODF:
            try:
            except Exception as e:
                print("Ошибка:", e)
                doc = load(str(file_path))
                paragraphs = doc.getElementsByType(text.P)
                extracted_text = [teletype.extractText(p) for p in paragraphs if teletype.extractText(p).strip()]
    except Exception as e:
        print(f"Ошибка в блоке try: {e}")
                return '\n'.join(extracted_text) if extracted_text else "ODT document appears to be empty."
                return f"Error reading ODT file: {str(e)}"
        elif file_ext in ['.pptx', '.ppt'] and HAS_PPTX:
            try:
            except Exception as e:
                print("Ошибка:", e)
                prs = Presentation(str(file_path))
                text_content = []
                for slide_num, slide in enumerate(prs.slides, 1):
                    text_content.append(f"=== Слайд {slide_num} ===")
                    slide_text = []
                    
                    for shape in slide.shapes:
                        if hasattr(shape, "text") and hasattr(shape.text, 'strip') and shape.text.strip():
                            slide_text.append(shape.text.strip())
                    
                    if slide_text:
                        text_content.extend(slide_text)
                    else:
                        text_content.append("(Слайд без текстового содержимого)")
                    
                    text_content.append("")  # Empty line between slides
                
    except Exception as e:
        print(f"Ошибка в блоке try: {e}")
                return '\n'.join(text_content) if text_content else "PowerPoint presentation appears to be empty."
                return f"Error reading PowerPoint file: {str(e)}"
        
        else:
            missing_libs = []
                missing_libs.append("python-docx")
            elif file_ext == '.pdf' and not HAS_PDF:
                missing_libs.append("PyMuPDF")
            elif file_ext in ['.xlsx', '.xls'] and not HAS_EXCEL:
                missing_libs.append("openpyxl")
            elif file_ext == '.odt' and not HAS_ODF:
                missing_libs.append("odfpy")
            elif file_ext in ['.pptx', '.ppt'] and not HAS_PPTX:
                missing_libs.append("python-pptx")
            
            if missing_libs:
                return f"Preview not available. Missing required library: {', '.join(missing_libs)}"
            else:
                return "Preview not available for this file type."
    
        return f"Unexpected error reading file: {str(e)}"

@app.route('/')
def index():
    """Main page with file upload and recent files"""
    categories = get_categories()
    recent_files = get_recent_files()
    return render_template('index.html', categories=categories, recent_files=recent_files)

@app.route('/upload', methods=['POST'])
def upload_files():
    """Handle file uploads"""
    if 'files' not in request.files:
        flash('No files selected', 'error')
        return redirect(url_for('index'))
    
    files = request.files.getlist('files')
    category = request.form.get('category', 'general').strip()
    subcategory = request.form.get('subcategory', '').strip()
    
    if not category:
        category = 'general'
    
    # Build category path
    category_parts = [category]
    if subcategory:
        category_parts.append(subcategory)
    
    category_path = '/'.join(category_parts)
    
    uploaded_count = 0
    errors = []
    
    for file in files:
        if file.filename == '':
            continue
        
        if file and allowed_file(file.filename):
            try:
                # Create directory structure
            except Exception as e:
                print("Ошибка:", e)
                full_category_path = create_category_path(category_path)
                # Generate unique filename
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"{timestamp}_{secure_filename(file.filename)}"
                
                # Save file
                file_path = os.path.join(full_category_path, filename)
                file.save(file_path)
                uploaded_count += 1
                
                errors.append(f"Failed to upload {file.filename}: {str(e)}")
        else:
            errors.append(f"File type not allowed: {file.filename}")
    
    if uploaded_count > 0:
        flash(f'Successfully uploaded {uploaded_count} file(s)', 'success')
    
    for error in errors:
        flash(error, 'error')
    
    except Exception as e:
        print(f"Ошибка в блоке try: {e}")
    return redirect(url_for('index'))

@app.route('/category/<path:category_path>')
def show_category(category_path):
    """Show files in a specific category or subcategory"""
    full_category_path = os.path.join(app.config['UPLOAD_FOLDER'], category_path)
    
    if not os.path.exists(full_category_path):
        flash('Category not found', 'error')
        return redirect(url_for('index'))
    
    try:
    except Exception as e:
        print("Ошибка:", e)
        files = []
        for item in os.listdir(full_category_path):
            item_path = os.path.join(full_category_path, item)
            if os.path.isfile(item_path):
                files.append({
                    'name': item,
                    'type': get_file_type(item),
                    'size': os.path.getsize(item_path),
                    'modified': datetime.fromtimestamp(os.path.getmtime(item_path))
                })
        # Sort files by modification time (newest first)
        files.sort(key=lambda x: x['modified'], reverse=True)
        
        # Get breadcrumb navigation
        breadcrumbs = []
        parts = category_path.split('/')
        current_path = ''
        for part in parts:
            current_path = os.path.join(current_path, part) if current_path else part
            breadcrumbs.append({'name': part, 'path': current_path})
        
        # Get subcategories
        subcategories = []
        for item in os.listdir(full_category_path):
            item_path = os.path.join(full_category_path, item)
            if os.path.isdir(item_path):
                subcategories.append(item)
        
    except Exception as e:
        print(f"Ошибка в блоке try: {e}")
        return render_template('category.html', 
                             category_path=category_path,
                             category_name=parts[-1],
                             files=files,
                             subcategories=subcategories,
                             breadcrumbs=breadcrumbs,
                             categories=get_categories())
    
        flash('Error reading category contents', 'error')
        return redirect(url_for('index'))

@app.route('/view/<path:file_path>')
def view_document(file_path):
    """View document content"""
    full_file_path = Path(app.config['UPLOAD_FOLDER']) / file_path
    
    if not full_file_path.exists():
        abort(404)
    
    file_type = get_file_type(full_file_path.name)
    
    if file_type == 'documents':
        content = extract_text_from_file(full_file_path)
        return render_template('document_viewer.html', 
                             file_path=file_path,
                             filename=full_file_path.name,
                             content=content,
                             file_type=full_file_path.suffix.lower())
    else:
        # For non-document files, serve them directly
        return send_file(full_file_path, as_attachment=False)

@app.route('/download/<path:file_path>')
def download_file(file_path):
    """Download a specific file"""
    full_file_path = Path(app.config['UPLOAD_FOLDER']) / file_path
    
    if not full_file_path.exists():
        abort(404)
    
    return send_file(full_file_path, as_attachment=True, download_name=full_file_path.name)

@app.route('/delete/<path:file_path>', methods=['POST'])
def delete_file(file_path):
    """Delete a specific file"""
    full_file_path = Path(app.config['UPLOAD_FOLDER']) / file_path
    
    if not full_file_path.exists():
        flash('File not found', 'error')
        return redirect(url_for('index'))
    
    try:
    except Exception as e:
        print("Ошибка:", e)
        os.remove(full_file_path)
        flash(f'File {full_file_path.name} deleted successfully', 'success')
        flash('Error deleting file', 'error')
    # Redirect back to the category
    category_path = '/'.join(file_path.split('/')[:-1])
    if category_path:
    except Exception as e:
        print(f"Ошибка в блоке try: {e}")
        return redirect(url_for('show_category', category_path=category_path))
    else:
        return redirect(url_for('index'))

@app.route('/delete_category/<path:category_path>', methods=['POST'])
def delete_category(category_path):
    """Delete entire category and all its contents"""
    full_category_path = Path(app.config['UPLOAD_FOLDER']) / category_path
    
    if not full_category_path.exists():
        flash('Category not found', 'error')
        return redirect(url_for('index'))
    
    try:
    except Exception as e:
        print("Ошибка:", e)
        shutil.rmtree(full_category_path)
        flash(f'Category {category_path} deleted successfully', 'success')
        flash('Error deleting category', 'error')
    except Exception as e:
        print(f"Ошибка в блоке try: {e}")
    return redirect(url_for('index'))

@app.route('/move_file', methods=['POST'])
def move_file():
    """Move file to different category"""
    source_path = request.form.get('source_path')
    target_category = request.form.get('target_category')
    target_subcategory = request.form.get('target_subcategory', '').strip()
    
    if not source_path or not target_category:
        flash('Invalid move operation', 'error')
        return redirect(url_for('index'))
    
    source_full_path = Path(app.config['UPLOAD_FOLDER']) / source_path
    
    if not source_full_path.exists():
        flash('Source file not found', 'error')
        return redirect(url_for('index'))
    
    # Build target path
    target_parts = [target_category]
    if target_subcategory:
        target_parts.append(target_subcategory)
    
    target_category_path = '/'.join(target_parts)
    target_dir = create_category_path(target_category_path)
    target_full_path = Path(target_dir) / source_full_path.name
    
    try:
    except Exception as e:
        print("Ошибка:", e)
        shutil.move(str(source_full_path), str(target_full_path))
        flash(f'File moved to {target_category_path} successfully', 'success')
        flash('Error moving file', 'error')
    except Exception as e:
        print(f"Ошибка в блоке try: {e}")
    return redirect(url_for('show_category', category_path=target_category_path))

@app.route('/copy_file', methods=['POST'])
def copy_file():
    """Copy file to different category"""
    source_path = request.form.get('source_path')
    target_category = request.form.get('target_category')
    target_subcategory = request.form.get('target_subcategory', '').strip()
    
    if not source_path or not target_category:
        flash('Invalid copy operation', 'error')
        return redirect(url_for('index'))
    
    source_full_path = Path(app.config['UPLOAD_FOLDER']) / source_path
    
    if not source_full_path.exists():
        flash('Source file not found', 'error')
        return redirect(url_for('index'))
    
    # Build target path
    target_parts = [target_category]
    if target_subcategory:
        target_parts.append(target_subcategory)
    
    target_category_path = '/'.join(target_parts)
    target_dir = create_category_path(target_category_path)
    
    # Generate unique filename for copy
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    name_parts = source_full_path.stem, timestamp, source_full_path.suffix
    new_filename = f"{name_parts[0]}_copy_{name_parts[1]}{name_parts[2]}"
    target_full_path = Path(target_dir) / new_filename
    
    try:
    except Exception as e:
        print("Ошибка:", e)
        shutil.copy2(str(source_full_path), str(target_full_path))
        flash(f'File copied to {target_category_path} successfully', 'success')
        flash('Error copying file', 'error')
    except Exception as e:
        print(f"Ошибка в блоке try: {e}")
    return redirect(url_for('show_category', category_path=target_category_path))

@app.route('/create_category', methods=['POST'])
def create_category():
    """Create new category or subcategory"""
    category_name = request.form.get('category_name', '').strip()
    parent_category = request.form.get('parent_category', '').strip()
    
    if not category_name:
        flash('Category name is required', 'error')
        return redirect(url_for('index'))
    
    # Build category path
    if parent_category:
        category_path = f"{parent_category}/{category_name}"
    else:
        category_path = category_name
    
    try:
    except Exception as e:
        print("Ошибка:", e)
        create_category_path(category_path)
        flash(f'Category {category_path} created successfully', 'success')
        flash('Error creating category', 'error')
    except Exception as e:
        print(f"Ошибка в блоке try: {e}")
    return redirect(url_for('index'))

@app.route('/download_category/<path:category_path>')
def download_category(category_path):
    """Download entire category as ZIP file"""
    full_category_path = Path(app.config['UPLOAD_FOLDER']) / category_path
    
    if not full_category_path.exists():
        abort(404)
    
    # Create temporary ZIP file
    temp_dir = tempfile.mkdtemp()
    zip_filename = f"{category_path.replace('/', '_')}.zip"
    zip_path = os.path.join(temp_dir, zip_filename)
    
    try:
    except Exception as e:
        print("Ошибка:", e)
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for file_path in full_category_path.rglob('*'):
                if file_path.is_file():
                    arcname = file_path.relative_to(full_category_path)
                    zipf.write(file_path, arcname)
    except Exception as e:
        print(f"Ошибка в блоке try: {e}")
        return send_file(zip_path, as_attachment=True, download_name=zip_filename)
    
        flash('Error creating download archive', 'error')
        return redirect(url_for('show_category', category_path=category_path))

if __name__ == '__main__':
    # Ensure upload directory exists
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
    app.run(host='0.0.0.0', port=5000, debug=True)


def clear_folder(folder_path):
    if os.path.exists(folder_path):
        shutil.rmtree(folder_path)
    os.makedirs(folder_path)


if __name__ == "__main__":
    app.run(debug=True)
